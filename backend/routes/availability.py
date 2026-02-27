import logging
import re
from datetime import datetime
from zoneinfo import ZoneInfo
from flask import Blueprint, request, jsonify
from dateutil import parser
import openpyxl
from extensions import db
from models import Worker
from config import DASH_PATTERN

availability_bp = Blueprint("availability_bp", __name__)


def to_time(val):
    from datetime import time as _time
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, _time):
        return val
    if isinstance(val, str):
        s = val.strip().upper()
        try:
            return datetime.strptime(s, "%H:%M").time()
        except ValueError:
            pass
        try:
            return datetime.strptime(s, "%I:%M %p").time()
        except ValueError:
            return None
    return None


@availability_bp.route('/upload-worker-availability', methods=['POST'])
def upload_worker_availability():
    try:
        file = request.files['file']
        if not file:
            return jsonify({'error': 'No file provided'}), 400

        workbook = openpyxl.load_workbook(file)

        def parse_time_range(cell_val):
            """
            Accepts strings like '08:00 - 16:00', '08:00–16:00', '8:00 AM — 4:30 PM'.
            Returns (start_time, end_time) as datetime.time.
            """
            s = str(cell_val).strip()

            m = re.search(rf"(\d{{1,2}}:\d{{2}})\s*{DASH_PATTERN}\s*(\d{{1,2}}:\d{{2}})", s)
            if m:
                t1 = datetime.strptime(m.group(1), "%H:%M").time()
                t2 = datetime.strptime(m.group(2), "%H:%M").time()
                return t1, t2

            m = re.search(rf"(\d{{1,2}}:\d{{2}}\s*[APap][Mm])\s*{DASH_PATTERN}\s*(\d{{1,2}}:\d{{2}}\s*[APap][Mm])", s)
            if m:
                t1 = datetime.strptime(m.group(1).upper(), "%I:%M %p").time()
                t2 = datetime.strptime(m.group(2).upper(), "%I:%M %p").time()
                return t1, t2

            raise ValueError(f"Unrecognized time range: {cell_val!r}")

        all_results = []  # collects a summary across all sheets
        updated_count = 0  # counter of DB updates

        # Process ALL worksheets in the file
        for sheet in workbook.worksheets:
            logging.info(f"🔎 Processing sheet: {sheet.title}")

            # Step 1: Extract date from B22 area on this sheet
            target_date = None
            for row in sheet.iter_rows(min_row=22, max_row=22):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and re.search(r"\d{2}/\d{2}/\d{4}", cell.value):
                        match = re.search(r"\d{2}/\d{2}/\d{4}", cell.value)
                        if match:
                            target_date = datetime.strptime(match.group(), "%d/%m/%Y")
                            break
                if target_date:
                    break

            if not target_date:
                logging.warning(f" Skipping sheet '{sheet.title}' — no date found on row 22.")
                continue  # move to next sheet

            # Step 2: Collect names and times from row 24 down on this sheet
            for i, row in enumerate(sheet.iter_rows(min_row=24), start=24):
                name_cell = row[1] if len(row) > 1 else None  # Column B
                time_cell = None

                # Check columns D, E, F for a single-cell range first
                for idx in [3, 4, 5]:
                    if len(row) > idx and row[idx].value:
                        time_cell = row[idx]
                        break

                logging.info(f"[{sheet.title}] Row {i} -> name: {name_cell.value if name_cell else 'None'}, time: {time_cell.value if time_cell else 'None'}")

                if not (name_cell and name_cell.value):
                    continue

                worker_name = str(name_cell.value).strip()

                # Try single-cell time range first
                start_t = end_t = None
                if time_cell and time_cell.value:
                    try:
                        start_t, end_t = parse_time_range(time_cell.value)
                    except Exception as parse_err:
                        logging.debug(f"[{sheet.title}] Single-cell time parse failed at row {i}: {parse_err}")

                # Fallback: if range not found in one cell, try separate start/end in D and E
                if (start_t is None or end_t is None):
                    start_cell = row[3] if len(row) > 3 else None  # col D
                    end_cell   = row[4] if len(row) > 4 else None  # col E

                    if start_cell and end_cell:
                        start_t = to_time(start_cell.value)
                        end_t = to_time(end_cell.value)

                if not (start_t and end_t):
                    # Nothing parseable on this row; continue to next row
                    continue

                # Record for response logging (keeps your existing behavior)
                time_range_display = f"{start_t.strftime('%H:%M')} - {end_t.strftime('%H:%M')}"
                all_results.append({"sheet": sheet.title, "name": worker_name, "time": time_range_display, "date": target_date.strftime("%Y-%m-%d")})

                # Update each worker's availability in the database (per-sheet date)
                existing_worker = Worker.query.filter_by(name=worker_name).first()
                if existing_worker:
                    try:
                        # Build datetimes on sheet's target_date using Europe/London timezone
                        start_datetime = datetime.combine(target_date.date(), start_t).replace(tzinfo=ZoneInfo("Europe/London"))
                        end_datetime   = datetime.combine(target_date.date(), end_t).replace(tzinfo=ZoneInfo("Europe/London"))

                        new_availability = {
                            "start": start_datetime.isoformat(),
                            "end": end_datetime.isoformat(),
                            "late": False
                        }

                        # Remove existing availability for this date (if any), then append
                        updated_availability = [
                            a for a in existing_worker.availability
                            if parser.parse(a["start"]).date() != target_date.date()
                        ]
                        updated_availability.append(new_availability)
                        existing_worker.availability = updated_availability

                        updated_count += 1
                    except Exception as parse_err:
                        logging.warning(f" Could not save time for {worker_name} (sheet '{sheet.title}', row {i}): {parse_err}")
                else:
                    logging.warning(f" Worker not found in DB: {worker_name}")

        # Commit once after processing all sheets (reduces I/O)
        db.session.commit()

        # Log the parsed availability (across all sheets)
        logging.info(" Parsed worker availability from Excel (all sheets):")
        for entry in all_results:
            logging.info(f"[{entry['sheet']}] {entry['date']} — {entry['name']} - {entry['time']}")
        logging.info(f"Total availability updates: {updated_count}")

        # Build response summary by date/sheet
        return jsonify({
            "updates": updated_count,
            "entries": all_results
        }), 200

    except Exception as e:
        logging.error(f" Error parsing availability upload: {e}")
        return jsonify({'error': str(e)}), 500
